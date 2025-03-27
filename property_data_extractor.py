#!/usr/bin/env python3
"""
Property Intelligence Data Extraction Tool

This script extracts property data from Excel files and imports it into a PostgreSQL database.
It handles work records, income, and expenses for multiple properties across different sheets.

Features:
- Automatically detects property addresses, dates, amounts, and descriptions
- Geocodes property addresses for mapping
- Prevents duplicate records when reimporting the same data
- Handles various date formats and Excel number formats
- Skips summary rows and headers (TOTAL, PROFIT, MARGIN, etc.)
- Supports existing database schemas
- Comprehensive error handling and reporting
"""

import pandas as pd
import psycopg2
from geopy.geocoders import Nominatim
import openpyxl
import datetime
import re
import time
import logging
import os
import sys
import argparse
import configparser
from typing import Dict, List, Tuple, Optional, Any, Union

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("propintel_import.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger()

# Database connection parameters
def get_db_config():
    """
    Reads database configuration from:
    1. DATABASE_URL environment variable (Heroku)
    2. db_config.ini file
    3. Default values as a last resort
    """
    # Check for DATABASE_URL environment variable (for Heroku)
    database_url = os.environ.get('DATABASE_URL')
    
    if database_url:
        logger.info("Using DATABASE_URL environment variable for connection")
        # Convert postgres:// to postgresql:// if necessary for psycopg2
        if database_url.startswith('postgres://'):
            database_url = database_url.replace('postgres://', 'postgresql://', 1)
        
        # Return as a DSN string
        return {'dsn': database_url}
    
    # Default connection parameters (only used if neither DATABASE_URL nor config file is available)
    default_params = {
        "user": "prop_intel",
        "password": "nyrty7-cytrit-qePkyf",
        "host": "propintel.postgres.database.azure.com",
        "port": "5432",
        "database": "postgres",
    }
    
    # Try to read from config file
    if os.path.exists('db_config.ini'):
        try:
            config = configparser.ConfigParser()
            config.read('db_config.ini')
            if 'database' in config:
                params = {
                    "user": config['database'].get('user', default_params['user']),
                    "password": config['database'].get('password', default_params['password']),
                    "host": config['database'].get('host', default_params['host']),
                    "port": int(config['database'].get('port', default_params['port'])),
                    "database": config['database'].get('database', default_params['database']),
                }
                logger.info("Using database configuration from db_config.ini")
                return params
        except Exception as e:
            logger.warning(f"Error reading config file: {e}. Using default parameters.")
    
    logger.info("Using default database parameters")
    # Use default parameters if file doesn't exist or has errors
    return default_params

# Get the connection parameters once at module level
connection_params = get_db_config()

def is_valid_date(date_value):
    """Check if a value is a valid date and not a header or summary row"""
    if date_value is None:
        return False
        
    # Check if it's already a date object
    if isinstance(date_value, datetime.datetime) or isinstance(date_value, datetime.date):
        return True
        
    # Skip common summary row labels
    if isinstance(date_value, str):
        skip_terms = ['total', 'totals', 'profit', 'margin', 'sum', 'average', 'header']
        if any(term in date_value.lower() for term in skip_terms):
            logger.debug(f"Skipping summary row with date value: {date_value}")
            return False
            
        # Try to parse common date formats
        try:
            if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', date_value):
                return True
            elif re.match(r'^\d{2,4}-\d{1,2}-\d{1,2}$', date_value):
                return True
        except:
            pass
            
    # Handle Excel numeric dates
    if isinstance(date_value, (int, float)) and 5000 < date_value < 50000:  # Reasonable date range
        try:
            # Test if it can be converted to a date
            datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(date_value))
            return True
        except:
            pass
            
    return False

def format_date(date_value):
    """Format a date value to a standard format, with validation"""
    if not is_valid_date(date_value):
        return None
        
    try:
        if isinstance(date_value, str):
            # Try different date formats
            if re.match(r'^\d{1,2}/\d{1,2}/\d{2,4}$', date_value):
                return datetime.datetime.strptime(date_value, '%d/%m/%y').date()
            elif re.match(r'^\d{2,4}-\d{1,2}-\d{1,2}$', date_value):
                return datetime.datetime.strptime(date_value, '%Y-%m-%d').date()
            else:
                logger.warning(f"Unrecognized date format: {date_value}")
                return None
        elif isinstance(date_value, (int, float)):
            # Excel dates are stored as days since 1900-01-01
            return (datetime.datetime(1899, 12, 30) + datetime.timedelta(days=int(date_value))).date()
        elif isinstance(date_value, (datetime.datetime, datetime.date)):
            # Already a date or datetime object
            if isinstance(date_value, datetime.datetime):
                return date_value.date()
            return date_value
    except Exception as e:
        logger.warning(f"Could not convert to date: {date_value}, Error: {e}")
        
    return None

def should_skip_row(row_data):
    """Check if a row should be skipped (e.g., summary rows, headers)"""
    # Skip rows with certain keywords in the date or description
    skip_terms = ['total', 'totals', 'profit', 'margin', 'sum', 'average', 'header']
    
    # Check date
    if isinstance(row_data['date'], str) and any(term in row_data['date'].lower() for term in skip_terms):
        return True
        
    # Check description
    if isinstance(row_data['description'], str) and any(term in row_data['description'].lower() for term in skip_terms):
        return True
    
    # Skip if date is invalid
    if not is_valid_date(row_data['date']):
        return True
    
    return False

def create_schema_and_tables(conn):
    """
    Creates the 'propintel' schema and necessary tables if they don't exist.
    Then checks and adds any missing columns to existing tables.
    """
    # First create schema and base tables
    create_sql = """
    -- Create schema if not exists
    CREATE SCHEMA IF NOT EXISTS propintel;

    -- Create tables within the 'propintel' schema
    CREATE TABLE IF NOT EXISTS propintel.properties (
        property_id SERIAL PRIMARY KEY,
        property_name VARCHAR(255),
        address TEXT,
        latitude DOUBLE PRECISION,
        longitude DOUBLE PRECISION
    );

    CREATE TABLE IF NOT EXISTS propintel.work (
        work_id SERIAL PRIMARY KEY,
        property_id INT NOT NULL,
        work_description TEXT,
        work_date DATE,
        work_cost NUMERIC(10, 2),
        FOREIGN KEY (property_id) REFERENCES propintel.properties (property_id)
    );

    CREATE TABLE IF NOT EXISTS propintel.money_in (
        money_in_id SERIAL PRIMARY KEY,
        property_id INT NOT NULL,
        income_amount NUMERIC(10,2),
        income_date DATE,
        income_details TEXT,
        FOREIGN KEY (property_id) REFERENCES propintel.properties (property_id)
    );

    CREATE TABLE IF NOT EXISTS propintel.money_out (
        money_out_id SERIAL PRIMARY KEY,
        property_id INT NOT NULL,
        expense_amount NUMERIC(10,2),
        expense_date DATE,
        expense_details TEXT,
        FOREIGN KEY (property_id) REFERENCES propintel.properties (property_id)
    );
    """
    
    try:
        with conn.cursor() as cur:
            cur.execute(create_sql)
            conn.commit()
        logger.info("Schema and base tables created or already exist")
        
        # Now check and add any missing columns
        add_columns_sql = []
        
        # Check if payment_method column exists in work table
        with conn.cursor() as cur:
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'propintel' 
                AND table_name = 'work'
                AND column_name = 'payment_method'
            """)
            if not cur.fetchone():
                add_columns_sql.append("""
                    ALTER TABLE propintel.work 
                    ADD COLUMN payment_method VARCHAR(50)
                """)
                logger.info("Will add payment_method column to work table")
            
            # Check if payment_method column exists in money_in table
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'propintel' 
                AND table_name = 'money_in'
                AND column_name = 'payment_method'
            """)
            if not cur.fetchone():
                add_columns_sql.append("""
                    ALTER TABLE propintel.money_in 
                    ADD COLUMN payment_method VARCHAR(50)
                """)
                logger.info("Will add payment_method column to money_in table")
            
            # Check if payment_method column exists in money_out table
            cur.execute("""
                SELECT column_name 
                FROM information_schema.columns 
                WHERE table_schema = 'propintel' 
                AND table_name = 'money_out'
                AND column_name = 'payment_method'
            """)
            if not cur.fetchone():
                add_columns_sql.append("""
                    ALTER TABLE propintel.money_out 
                    ADD COLUMN payment_method VARCHAR(50)
                """)
                logger.info("Will add payment_method column to money_out table")
        
        # Apply any needed column additions
        if add_columns_sql:
            with conn.cursor() as cur:
                for sql in add_columns_sql:
                    try:
                        cur.execute(sql)
                        conn.commit()
                        logger.info("Successfully added column")
                    except Exception as e:
                        logger.error(f"Error adding column: {e}")
                        conn.rollback()
        
    except Exception as e:
        logger.error(f"Error creating schema and tables: {e}")
        conn.rollback()
        raise

def geocode_address(address, geolocator, max_retries=3, retry_delay=1):
    """
    Attempt to geocode the address using geopy with retry logic.
    Returns (latitude, longitude) or (None, None) if not found or on error.
    """
    if not address or not isinstance(address, str):
        logger.warning(f"Invalid address: {address}")
        return (None, None)
    
    # Clean up the address
    address = address.strip()
    if not address:
        return (None, None)
    
    # Add Australia suffix if not present
    if "australia" not in address.lower():
        address = f"{address}, Australia"
    
    for attempt in range(max_retries):
        try:
            logger.info(f"Geocoding address: {address}")
            location = geolocator.geocode(address)
            if location:
                logger.info(f"Geocoding succeeded: {location.latitude}, {location.longitude}")
                return (location.latitude, location.longitude)
            else:
                logger.warning(f"No geocoding results for address: {address}")
        except Exception as e:
            logger.warning(f"Geocoding error on attempt {attempt+1} for address '{address}': {e}")
        
        if attempt < max_retries - 1:
            logger.info(f"Retrying in {retry_delay} seconds...")
            time.sleep(retry_delay)
    
    return (None, None)

def get_property_id(conn, property_address, property_name=None):
    """
    Check if a property already exists (by address). If not found, insert it.
    Returns the property_id.
    """
    if not property_address:
        raise ValueError("Property address is required")
    
    # Normalize address
    address_cleaned = property_address.strip()
    
    # Default property name if not provided
    if not property_name:
        property_name = f"Property at {address_cleaned}"
    
    try:
        with conn.cursor() as cur:
            # Check if property exists by address
            check_sql = """
                SELECT property_id 
                FROM propintel.properties
                WHERE LOWER(address) = LOWER(%s)
            """
            cur.execute(check_sql, (address_cleaned,))
            result = cur.fetchone()

            if result:
                # Property already exists
                logger.info(f"Found existing property ID {result[0]} for address: {address_cleaned}")
                return result[0]
            else:
                # Initialize geolocator for new property
                geolocator = Nominatim(user_agent="propintel_geocoder")
                
                # Geocode the address
                latitude, longitude = geocode_address(address_cleaned, geolocator)
                
                # Insert new property
                insert_sql = """
                    INSERT INTO propintel.properties (property_name, address, latitude, longitude)
                    VALUES (%s, %s, %s, %s)
                    RETURNING property_id
                """
                cur.execute(insert_sql, (property_name, address_cleaned, latitude, longitude))
                new_id = cur.fetchone()[0]
                conn.commit()
                logger.info(f"Created new property with ID {new_id} for address: {address_cleaned}")
                return new_id
    except Exception as e:
        logger.error(f"Error getting/creating property ID for {address_cleaned}: {e}")
        conn.rollback()
        raise

def insert_work(conn, property_id, work_description, work_date, work_cost, payment_method=None):
    """Insert work record with duplicate checking and validation"""
    if not property_id or not work_description:
        logger.warning("Missing required fields for work record")
        return False
    
    # Validate and clean data
    try:
        # Format date
        work_date = format_date(work_date)
        
        # Format cost
        if work_cost is not None:
            try:
                work_cost = float(work_cost)
            except:
                logger.warning(f"Invalid work cost: {work_cost}, using None")
                work_cost = None
        
        # Skip if missing key data
        if work_date is None:
            logger.warning(f"Skipping work record with invalid date for property {property_id}")
            return False
        
        with conn.cursor() as cur:
            # First check if a similar record already exists
            similar_record_sql = """
                SELECT work_id 
                FROM propintel.work 
                WHERE property_id = %s 
                AND work_date = %s 
                AND TRIM(work_description) = TRIM(%s)
            """
            
            cur.execute(similar_record_sql, (property_id, work_date, work_description))
            existing_record = cur.fetchone()
            
            if existing_record:
                # Record exists - update it if needed
                work_id = existing_record[0]
                
                # Check if payment_method column exists
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'propintel' 
                    AND table_name = 'work'
                    AND column_name = 'payment_method'
                """)
                
                if cur.fetchone():
                    # Update with payment_method
                    update_sql = """
                        UPDATE propintel.work 
                        SET work_cost = %s, payment_method = %s
                        WHERE work_id = %s
                    """
                    cur.execute(update_sql, (work_cost, payment_method, work_id))
                else:
                    # Update without payment_method
                    update_sql = """
                        UPDATE propintel.work 
                        SET work_cost = %s
                        WHERE work_id = %s
                    """
                    cur.execute(update_sql, (work_cost, work_id))
                
                conn.commit()
                logger.info(f"Updated existing work record {work_id} for property {property_id}")
                return True
            else:
                # No existing record - insert new one
                # Check if payment_method column exists
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'propintel' 
                    AND table_name = 'work'
                    AND column_name = 'payment_method'
                """)
                
                if cur.fetchone():
                    # Use payment_method if column exists
                    sql = """
                        INSERT INTO propintel.work (property_id, work_description, work_date, work_cost, payment_method)
                        VALUES (%s, %s, %s, %s, %s)
                    """
                    cur.execute(sql, (property_id, work_description, work_date, work_cost, payment_method))
                else:
                    # Skip payment_method if column doesn't exist
                    sql = """
                        INSERT INTO propintel.work (property_id, work_description, work_date, work_cost)
                        VALUES (%s, %s, %s, %s)
                    """
                    cur.execute(sql, (property_id, work_description, work_date, work_cost))
                
                conn.commit()
                logger.info(f"Inserted new work record for property {property_id}: {work_description[:30]}..., {work_date}, ${work_cost}")
                return True
    except Exception as e:
        logger.error(f"Error processing work record: {e}")
        conn.rollback()
        return False

def insert_money_in(conn, property_id, amount, date, details, payment_method=None):
    """Insert money in record with duplicate checking and validation"""
    if not property_id or amount is None:
        logger.warning("Missing required fields for money in record")
        return False
    
    try:
        # Format date
        date = format_date(date)
        
        # Format amount
        if amount is not None:
            try:
                amount = float(amount)
            except:
                logger.warning(f"Invalid amount: {amount}, using None")
                amount = None
        
        # Skip if missing key data
        if date is None:
            logger.warning(f"Skipping money in record with invalid date for property {property_id}")
            return False
        
        with conn.cursor() as cur:
            # First check if a similar record already exists
            similar_record_sql = """
                SELECT money_in_id 
                FROM propintel.money_in 
                WHERE property_id = %s 
                AND income_date = %s 
                AND ABS(income_amount - %s) < 0.01
            """
            
            cur.execute(similar_record_sql, (property_id, date, amount))
            existing_record = cur.fetchone()
            
            if existing_record:
                # Record exists - update it if needed
                money_in_id = existing_record[0]
                
                # Check if payment_method column exists
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'propintel' 
                    AND table_name = 'money_in'
                    AND column_name = 'payment_method'
                """)
                
                if cur.fetchone():
                    # Update with payment_method
                    update_sql = """
                        UPDATE propintel.money_in 
                        SET income_details = %s, payment_method = %s
                        WHERE money_in_id = %s
                    """
                    cur.execute(update_sql, (details, payment_method, money_in_id))
                else:
                    # Update without payment_method
                    update_sql = """
                        UPDATE propintel.money_in 
                        SET income_details = %s
                        WHERE money_in_id = %s
                    """
                    cur.execute(update_sql, (details, money_in_id))
                
                conn.commit()
                logger.info(f"Updated existing money in record {money_in_id} for property {property_id}")
                return True
            else:
                # No existing record - insert new one
                # Check if payment_method column exists
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'propintel' 
                    AND table_name = 'money_in'
                    AND column_name = 'payment_method'
                """)
                
                if cur.fetchone():
                    # Use payment_method if column exists
                    sql = """
                        INSERT INTO propintel.money_in (property_id, income_amount, income_date, income_details, payment_method)
                        VALUES (%s, %s, %s, %s, %s)
                    """
                    cur.execute(sql, (property_id, amount, date, details, payment_method))
                else:
                    # Skip payment_method if column doesn't exist
                    sql = """
                        INSERT INTO propintel.money_in (property_id, income_amount, income_date, income_details)
                        VALUES (%s, %s, %s, %s)
                    """
                    cur.execute(sql, (property_id, amount, date, details))
                
                conn.commit()
                logger.info(f"Inserted new money in record for property {property_id}: ${amount}, {date}")
                return True
    except Exception as e:
        logger.error(f"Error processing money in record: {e}")
        conn.rollback()
        return False

def insert_money_out(conn, property_id, amount, date, details, payment_method=None):
    """Insert money out record with duplicate checking and validation"""
    if not property_id or amount is None:
        logger.warning("Missing required fields for money out record")
        return False
    
    try:
        # Format date
        date = format_date(date)
        
        # Format amount
        if amount is not None:
            try:
                amount = float(amount)
            except:
                logger.warning(f"Invalid amount: {amount}, using None")
                amount = None
        
        # Skip if missing key data
        if date is None:
            logger.warning(f"Skipping money out record with invalid date for property {property_id}")
            return False
        
        with conn.cursor() as cur:
            # First check if a similar record already exists
            similar_record_sql = """
                SELECT money_out_id 
                FROM propintel.money_out 
                WHERE property_id = %s 
                AND expense_date = %s 
                AND ABS(expense_amount - %s) < 0.01
            """
            
            cur.execute(similar_record_sql, (property_id, date, amount))
            existing_record = cur.fetchone()
            
            if existing_record:
                # Record exists - update it if needed
                money_out_id = existing_record[0]
                
                # Check if payment_method column exists
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'propintel' 
                    AND table_name = 'money_out'
                    AND column_name = 'payment_method'
                """)
                
                if cur.fetchone():
                    # Update with payment_method
                    update_sql = """
                        UPDATE propintel.money_out 
                        SET expense_details = %s, payment_method = %s
                        WHERE money_out_id = %s
                    """
                    cur.execute(update_sql, (details, payment_method, money_out_id))
                else:
                    # Update without payment_method
                    update_sql = """
                        UPDATE propintel.money_out 
                        SET expense_details = %s
                        WHERE money_out_id = %s
                    """
                    cur.execute(update_sql, (details, money_out_id))
                
                conn.commit()
                logger.info(f"Updated existing money out record {money_out_id} for property {property_id}")
                return True
            else:
                # No existing record - insert new one
                # Check if payment_method column exists
                cur.execute("""
                    SELECT column_name 
                    FROM information_schema.columns 
                    WHERE table_schema = 'propintel' 
                    AND table_name = 'money_out'
                    AND column_name = 'payment_method'
                """)
                
                if cur.fetchone():
                    # Use payment_method if column exists
                    sql = """
                        INSERT INTO propintel.money_out (property_id, expense_amount, expense_date, expense_details, payment_method)
                        VALUES (%s, %s, %s, %s, %s)
                    """
                    cur.execute(sql, (property_id, amount, date, details, payment_method))
                else:
                    # Skip payment_method if column doesn't exist
                    sql = """
                        INSERT INTO propintel.money_out (property_id, expense_amount, expense_date, expense_details)
                        VALUES (%s, %s, %s, %s)
                    """
                    cur.execute(sql, (property_id, amount, date, details))
                
                conn.commit()
                logger.info(f"Inserted new money out record for property {property_id}: ${amount}, {date}")
                return True
    except Exception as e:
        logger.error(f"Error processing money out record: {e}")
        conn.rollback()
        return False

def extract_property_address(sheet):
    """Extract the property address from the sheet"""
    # Addresses are usually in the first cell of the second row
    try:
        if sheet.cell(row=2, column=2).value:
            address = str(sheet.cell(row=2, column=2).value).strip()
            if address and not address.startswith("TOTAL") and not address.lower() == "date":
                return address
        
        # If not found in the standard location, scan the first few rows
        for row in range(1, 10):
            for col in range(1, 5):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    # Look for text that seems like an address
                    address = cell_value.strip()
                    if re.search(r'\d+\s+\w+\s+(?:St|Ave|Rd|Dr|Crt|Cres|Street|Avenue|Road)', address, re.IGNORECASE):
                        return address
        
        # Default to sheet name if nothing else found
        sheet_name = sheet.title
        if not sheet_name.startswith("Sheet"):
            return f"Property: {sheet_name}"
        
        return None
    except Exception as e:
        logger.error(f"Error extracting property address: {e}")
        return None

def identify_columns(sheet):
    """
    Analyze the sheet to identify column mappings based on headers or patterns.
    Returns a dictionary mapping column numbers to data types.
    """
    column_map = {}
    header_row = None
    
    # Try to find a header row (within first 10 rows)
    for row_idx in range(1, min(10, sheet.max_row + 1)):
        header_candidates = []
        for col_idx in range(1, min(15, sheet.max_column + 1)):
            cell_value = sheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                header_candidates.append((col_idx, str(cell_value).strip().lower()))
        
        # If we have several header candidates, this might be our header row
        if len(header_candidates) >= 3:
            date_cols = []
            amount_cols = []
            desc_cols = []
            method_cols = []
            
            for col_idx, header in header_candidates:
                if header in ('date', 'work date', 'money in date', 'money out date'):
                    date_cols.append(col_idx)
                elif header in ('amount', 'cost', 'work cost', 'money in', 'money out', 'expense amount', 'blue ladder'):
                    amount_cols.append(col_idx)
                elif header in ('item', 'details', 'description', 'work description', 'expense details'):
                    desc_cols.append(col_idx)
                elif header in ('method', 'payment method', 'payment'):
                    method_cols.append(col_idx)
            
            # If we identified some columns, use this as our header row
            if date_cols or amount_cols:
                header_row = row_idx
                
                # Map columns
                for col_idx, header in header_candidates:
                    if 'date' in header:
                        column_map[col_idx] = 'date'
                    elif header in ('amount', 'cost', 'expense amount', 'income amount', 'money in', 'money out', 'blue ladder'):
                        column_map[col_idx] = 'amount'
                    elif header in ('item', 'details', 'description', 'work description', 'expense details', 'income details'):
                        column_map[col_idx] = 'description'
                    elif header in ('method', 'payment method', 'payment'):
                        column_map[col_idx] = 'method'
                
                # If we found good column mappings, break
                if column_map:
                    break
    
    # If no header row identified, make educated guesses based on content patterns
    if not column_map:
        # Check first data rows for patterns
        date_candidates = set()
        amount_candidates = set()
        desc_candidates = set()
        method_candidates = set()
        
        # Sample several rows to detect column types
        sample_rows = [11, 15, 20, 25, 30]
        valid_samples = 0
        
        for row_idx in sample_rows:
            if row_idx > sheet.max_row:
                continue
                
            valid_samples += 1
            for col_idx in range(1, min(15, sheet.max_column + 1)):
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if not cell_value:
                    continue
                
                # Check for dates
                if isinstance(cell_value, datetime.datetime):
                    date_candidates.add(col_idx)
                elif isinstance(cell_value, str) and re.match(r'\d{1,2}/\d{1,2}/\d{2,4}', cell_value):
                    date_candidates.add(col_idx)
                    
                # Check for amounts
                if isinstance(cell_value, (int, float)) and not isinstance(cell_value, bool):
                    amount_candidates.add(col_idx)
                
                # Check for method - usually short strings like CASH, TRANSFER
                if isinstance(cell_value, str) and cell_value.upper() in ('CASH', 'CARD', 'TRANSFER', 'CHECK', 'CHEQUE'):
                    method_candidates.add(col_idx)
                
                # Check for descriptions - longer strings
                if isinstance(cell_value, str) and len(cell_value) > 20:
                    desc_candidates.add(col_idx)
        
        # We need at least some valid samples
        if valid_samples >= 2:
            # Assign column types based on patterns
            header_row = 0  # no header row
            
            # First date column
            if date_candidates:
                date_col = min(date_candidates)
                column_map[date_col] = 'date'
            
            # Description columns
            for col in desc_candidates:
                column_map[col] = 'description'
            
            # Amount columns (excluding dates)
            for col in amount_candidates:
                if col not in column_map:
                    column_map[col] = 'amount'
            
            # Method columns
            for col in method_candidates:
                column_map[col] = 'method'
    
    return header_row, column_map

def process_sheet(conn, sheet):
    """Process a single sheet and extract property data"""
    logger.info(f"Processing sheet: {sheet.title}")
    
    # Extract property address from sheet
    property_address = extract_property_address(sheet)
    if not property_address:
        logger.warning(f"No property address found in sheet {sheet.title}, skipping")
        return 0
    
    logger.info(f"Found property address: {property_address}")
    
    # Get or create property ID
    try:
        property_id = get_property_id(conn, property_address)
    except Exception as e:
        logger.error(f"Could not get property ID for {property_address}: {e}")
        return 0
    
    # Identify column structure
    header_row, column_map = identify_columns(sheet)
    if not column_map:
        logger.warning(f"Could not identify column structure in sheet {sheet.title}, using default mappings")
        # Use default column mappings (common in the observed data)
        header_row = 4  # Headers often in row 5
        column_map = {
            2: 'date',        # Column B
            3: 'description',  # Column C
            4: 'amount',      # Column D
            6: 'method',      # Column F
            8: 'date',        # Column H (money in date)
            9: 'amount',      # Column I (money in amount)
            10: 'method'      # Column J (money in method)
        }
    
    # Start processing from the row after the header (or row 1 if no header)
    start_row = header_row + 1 if header_row else 2
    
    # Get columns by type
    date_cols = [col for col, type in column_map.items() if type == 'date']
    amount_cols = [col for col, type in column_map.items() if type == 'amount']
    desc_cols = [col for col, type in column_map.items() if type == 'description']
    method_cols = [col for col, type in column_map.items() if type == 'method']
    
    # Track records inserted
    records_inserted = 0
    records_updated = 0
    records_skipped = 0
    
    # Process each row
    for row_idx in range(start_row, sheet.max_row + 1):
        # Initialize data for this row
        row_data = {
            'date': None,
            'description': None,
            'amount': None,
            'method': None
        }
        
        # Check if row has any data
        has_data = False
        for col_idx in range(1, min(15, sheet.max_column + 1)):
            if sheet.cell(row=row_idx, column=col_idx).value is not None:
                has_data = True
                break
                
        if not has_data:
            continue
        
        # Extract data from this row
        for col_type, col_list in [
            ('date', date_cols),
            ('amount', amount_cols),
            ('description', desc_cols),
            ('method', method_cols)
        ]:
            for col_idx in col_list:
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    row_data[col_type] = cell_value
                    break
        
        # Skip summary rows
        if should_skip_row(row_data):
            records_skipped += 1
            continue
        
        # Decide if this is work, money in, or money out
        # For simplicity: 
        # - Positive amounts in cols 1-6 are work
        # - Positive amounts in cols 7+ are money in
        # - Negative amounts are money out
        
        amount = row_data['amount']
        if amount is None:
            records_skipped += 1
            continue
            
        amount_col = None
        for col_idx in amount_cols:
            if sheet.cell(row=row_idx, column=col_idx).value is not None:
                amount_col = col_idx
                break
        
        success = False
        if amount_col is not None:
            try:
                amount_value = float(amount)
                date_value = row_data['date']
                desc_value = row_data['description'] or f"Transaction on {date_value}"
                method_value = row_data['method']
                
                if amount_col <= 6:
                    # This is work or expense
                    if amount_value >= 0:
                        success = insert_work(conn, property_id, desc_value, date_value, amount_value, method_value)
                    else:
                        success = insert_money_out(conn, property_id, abs(amount_value), date_value, desc_value, method_value)
                else:
                    # This is likely income
                    success = insert_money_in(conn, property_id, amount_value, date_value, desc_value, method_value)
            except Exception as e:
                logger.warning(f"Error processing row {row_idx}: {e}")
                records_skipped += 1
                continue
        
        if success:
            records_inserted += 1
    
    logger.info(f"Processed {records_inserted + records_skipped} rows for property: {property_address}")
    logger.info(f"Inserted/updated {records_inserted} records, skipped {records_skipped} rows")
    return records_inserted

def clean_database(conn, confirm=False):
    """
    WARNING: This will delete all data from the database.
    Use only when you want to start fresh.
    """
    if not confirm:
        logger.warning("Database cleanup was requested but not confirmed. Skipping cleanup.")
        return False
    
    try:
        with conn.cursor() as cur:
            # Drop tables in correct order (respecting foreign key constraints)
            logger.info("Dropping all tables in propintel schema...")
            cur.execute("""
                DROP TABLE IF EXISTS propintel.money_in CASCADE;
                DROP TABLE IF EXISTS propintel.money_out CASCADE;
                DROP TABLE IF EXISTS propintel.work CASCADE;
                DROP TABLE IF EXISTS propintel.properties CASCADE;
            """)
            conn.commit()
            logger.info("All tables have been dropped.")
            return True
    except Exception as e:
        logger.error(f"Error during database cleanup: {e}")
        conn.rollback()
        return False

def extract_data_from_excel(file_path):
    """
    1) Connect to DB
    2) Create schema and tables if not exist
    3) Read Excel
    4) For each sheet, process data and insert into relevant tables
    5) Print progress and final summary
    """
    logger.info(f"Starting extraction process from file: {file_path}")

    # Check if file exists
    if not os.path.exists(file_path):
        logger.error(f"Excel file not found: {file_path}")
        return
    
    try:
        # Connect to PostgreSQL database
        conn = psycopg2.connect(**connection_params)
        logger.info("Connected to PostgreSQL database")
        
        # Create schema and tables if they don't exist
        create_schema_and_tables(conn)
        
        # Load the Excel workbook
        logger.info(f"Loading Excel file: {file_path}")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        
        # Count all sheets
        total_sheets = len(workbook.sheetnames)
        logger.info(f"Found {total_sheets} sheets in the Excel file")
        
        # Process each sheet
        total_records = 0
        sheets_processed = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Skip empty sheets
            if sheet.max_row < 2 or sheet.max_column < 2:
                logger.info(f"Skipping empty sheet: {sheet_name}")
                continue
            
            try:
                records = process_sheet(conn, sheet)
                total_records += records
                sheets_processed += 1
                logger.info(f"Processed sheet {sheets_processed}/{total_sheets}: {sheet_name}")
            except Exception as e:
                logger.error(f"Error processing sheet {sheet_name}: {e}")
                continue
        
        # Summary
        logger.info(f"Extraction completed. Processed {sheets_processed} sheets, inserted {total_records} records.")
        conn.close()
        logger.info("Database connection closed")
        
    except Exception as e:
        logger.error(f"Extraction failed: {e}")
        raise

def test_db_connection():
    """Test the database connection and return True if successful"""
    try:
        conn = psycopg2.connect(**connection_params)
        with conn.cursor() as cur:
            cur.execute("SELECT version();")
            version = cur.fetchone()
            print(f"Connected to PostgreSQL: {version[0]}")
        conn.close()
        return True
    except Exception as e:
        print(f"Database connection failed: {e}")
        return False

def main():
    # Command line arguments
    parser = argparse.ArgumentParser(description='Property Intelligence Data Import Tool')
    parser.add_argument('--excel', default="uploads/Master_Sheet.xlsx", help='Path to Excel file')
    parser.add_argument('--clean', action='store_true', help='Clean database before import (WARNING: deletes all data)')
    parser.add_argument('--test-connection', action='store_true', help='Test database connection and exit')
    args = parser.parse_args()
    
    # Test database connection if requested
    if args.test_connection:
        success = test_db_connection()
        sys.exit(0 if success else 1)
    
    # Check if clean is requested
    if args.clean:
        print("WARNING: You have requested to clean the database.")
        print("This will DELETE ALL DATA in the propintel schema.")
        confirmation = input("Type 'DELETE' to confirm: ")
        
        if confirmation == 'DELETE':
            try:
                conn = psycopg2.connect(**connection_params)
                clean_database(conn, confirm=True)
                conn.close()
                print("Database cleaned successfully.")
            except Exception as e:
                print(f"Error cleaning database: {e}")
                return
        else:
            print("Database cleanup cancelled.")
    
    # Check if Excel file exists
    if not os.path.exists(args.excel):
        print(f"ERROR: Excel file '{args.excel}' not found.")
        return
    
    # Proceed with data extraction
    extract_data_from_excel(args.excel)

if __name__ == "__main__":
    main()